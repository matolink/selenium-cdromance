import openpyxl
from openpyxl import Workbook
from program import pg
from check import ck
import waitdl as wd

excel = 'MENTIRITAS.xlsx'
DIR = '/home/matito/Downloads/'
CON = 'GBA'

class Game:
    path = DIR
    console = CON
    file_name = ''
    num = 1
    def __init__(self, name):
        self.name = name
    def program(self):
        pg(self)
    def wait(self):
        wd.waitdl(self)
    def check(self):
        ck(self)
        

if 'DIR' in locals():
    wb_obj = openpyxl.load_workbook(excel)
    sheet_obj = wb_obj.active
    max_row = sheet_obj.max_row
    workbook = Workbook()
    sheet = workbook.active
    nf = 1
    for i in range(2, max_row):
        GAME = Game(sheet_obj.cell(row=i, column=1))
        if GAME.name.value == None:
            exit()
        else:
            GAME.num = i - 1
            print(i-1, GAME.name.value)
            not_found = pg(GAME)
            if(not_found == True):
                sheet["A"+str(nf)] = str(GAME.num)+' '+GAME.name.value
                nf+=1
                workbook.save("not founds.xlsx")
                continue
else:
    print('DIR is not defined! please define a DIR and start the program again')
    exit()
